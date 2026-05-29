"""Parser dos relatórios W011A (despesas) e W045A (fração ideal) do Superlógica.

Lê os PDFs, identifica o tipo pelo conteúdo (não pelo nome), extrai os dados e
gera a planilha de previsão orçamentária no mesmo formato da referência
`skills-server/Previsao_Naturale_2026.xlsx`.

Uso CLI:
    python parser_superlogica.py --w011a w011a.pdf --w045a w045a.pdf --saida previsao.xlsx

Uso como módulo:
    from parser_superlogica import parsear_w011a, parsear_w045a, gerar_planilha
    d11 = parsear_w011a('w011a.pdf')
    d45 = parsear_w045a('w045a.pdf')
    gerar_planilha(d11, d45, 'previsao.xlsx')

Limitações conhecidas:
- O parsing é heurístico (regex + palavras-chave de categoria). Quando o
  layout do Superlógica mudar, atualizar `_PADROES_*` no topo deste arquivo.
- Subcategorias não previstas em CATEGORIA_KEYWORDS caem em "Outros" e geram
  warning no stderr para o operador mapear.
- Valores monetários no PDF assumem padrão BR (vírgula decimal, ponto milhar).
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from typing import Any

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─── Constantes canônicas (ondas 1 a 3 da prestação de contas) ───────────

CATEGORIAS_CANONICAS = [
    'Despesas Financeiras',
    'Despesa com Funcionários',
    'Retenções Fiscais',
    'Despesa Administrativa',
    'Manutenção',
    'Aquisição de Materiais',
    'Serviços',
    'Investimento e Equipamentos',
    'Taxas e Recolhimentos',
]

# Palavras-chave que mapeiam uma subcategoria/item do W011A para um dos 9 grupos.
# Match por substring lower-case na primeira ocorrência. Ordem matters.
CATEGORIA_KEYWORDS: dict[str, list[str]] = {
    'Despesas Financeiras': [
        'tarifa banc', 'tarifas banc', 'irrf poupan', 'iof', 'reembolso',
        'cotas de capital', 'devolução pagamento', 'empréstimo', 'emprestimo',
        'taxa antecipação', 'taxa antecipacao',
    ],
    'Despesa com Funcionários': [
        'mão de obra', 'mao de obra', 'salár', 'salar', 'encargo',
        'inss patronal', 'cesta básica', 'cesta basica', 'vale transporte',
        'alimentação', 'alimentacao', 'padaria',
    ],
    'Retenções Fiscais': [
        'iss', 'darf', 'irrf', 'pis cofins', 'csll', 'das simples',
    ],
    'Despesa Administrativa': [
        'honorário', 'honorario', 'cartório', 'cartorio', 'correios',
        'material de expediente', 'custas processuais', 'multa', 'penalidade',
        'deslocamento', 'comissão de cobrança', 'comissao de cobranca',
    ],
    'Manutenção': [
        'contrato manut', 'contrato cx.', 'manutenção elevador',
        'manutenção piscina', 'manutenção bombas', 'manutenção segurança',
        'manutenção jardinagem', 'manutenção desinsetizaç',
        'manutenção ete', 'manutenção segurança eletrôni',
        'cftv', 'fossa e esgo',
    ],
    'Aquisição de Materiais': [
        'material', 'tags', 'ferramentas', 'materiais obras',
    ],
    'Serviços': [
        'seguro condom', 'serviço', 'servico', 'obras-melhorias',
        'obras e melhorias', 'sistema de incêndio', 'sistema de incendio',
    ],
    'Investimento e Equipamentos': [
        'móvel', 'movel', 'utensílio', 'utensilio', 'informática',
        'informatica', 'eletrodom', 'máquina', 'maquina', 'equipamento',
        'container de lixo',
    ],
    'Taxas e Recolhimentos': [
        'art -', 'art–', 'conselho regional', 'alvará', 'alvara', 'iptu',
        'taxa', 'recolhimento',
    ],
}

# Itens que ficam FORA do rateio condominial. Rateados separadamente por uso
# real (Energia/Água) ou tratados como dívida extraordinária (Empréstimo).
PADROES_FORA_RATEIO = [
    'energia', 'água individual', 'agua individual', 'água-esgoto',
    'agua-esgoto', 'água e esgoto', 'agua e esgoto', 'telefone',
    'telefonia', 'internet', 'gás individual', 'gas individual',
    'empréstimo', 'emprestimo', 'obras extraordinárias',
    'obras extraordinarias', 'materiais obras-melhorias',
]

# ─── Identificação do tipo de PDF ────────────────────────────────────────

_MARCADORES_W011A = [
    'w011a', 'demonstrativo de despesas', 'despesas - últimos 12',
    'despesas dos ultimos 12', 'despesas mensais por categoria',
    'lançamentos do período',
]
_MARCADORES_W045A = [
    'w045a', 'fração ideal', 'fracao ideal', 'frações ideais',
    'fracoes ideais', 'rateio por fração', 'rateio por fracao',
]


def identificar_tipo_pdf(caminho: str) -> str:
    """Identifica se o PDF é W011A, W045A ou DESCONHECIDO pelo conteúdo.

    Lê só a primeira página por performance. Se a heurística não acertar
    pelo cabeçalho, varre as 3 primeiras páginas como fallback.
    """
    try:
        with pdfplumber.open(caminho) as pdf:
            paginas_consideradas = pdf.pages[:3]
            texto = '\n'.join(
                (p.extract_text() or '') for p in paginas_consideradas
            ).lower()
    except Exception as e:
        print(f'[parser] falha ao abrir {caminho}: {e}', file=sys.stderr)
        return 'DESCONHECIDO'

    if any(m in texto for m in _MARCADORES_W011A):
        return 'W011A'
    if any(m in texto for m in _MARCADORES_W045A):
        return 'W045A'
    return 'DESCONHECIDO'


# ─── Helpers de parsing ──────────────────────────────────────────────────

_RE_VALOR_BRL = re.compile(r'(-?\d{1,3}(?:\.\d{3})*,\d{2})')
_RE_DATA_BR = re.compile(r'(\d{2}/\d{2}/\d{4})')
_RE_PERIODO = re.compile(
    r'((?:jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)/\d{4})\s*(?:a|até|–|-)\s*'
    r'((?:jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)/\d{4})',
    re.IGNORECASE,
)


def _parse_valor(texto: str) -> float | None:
    """Converte string BR ("1.234,56" ou "1234,56") em float. None se inválido."""
    if not texto:
        return None
    s = texto.strip().replace(' ', '')
    m = _RE_VALOR_BRL.search(s)
    if not m:
        return None
    return float(m.group(1).replace('.', '').replace(',', '.'))


def _classificar_categoria(nome_item: str) -> str:
    """Retorna a categoria canônica de um item, ou 'Outros' se não casar."""
    s = (nome_item or '').lower()
    for cat, palavras in CATEGORIA_KEYWORDS.items():
        if any(p in s for p in palavras):
            return cat
    return 'Outros'


def _eh_fora_rateio(nome_item: str) -> bool:
    s = (nome_item or '').lower()
    return any(p in s for p in PADROES_FORA_RATEIO)


# ─── W011A ───────────────────────────────────────────────────────────────

def parsear_w011a(caminho: str) -> dict[str, Any]:
    """Extrai dados do W011A (demonstrativo de despesas, 12 meses).

    Retorna dict com:
        condominio:           str
        periodo:              str  (ex "Mai/2025 a Abr/2026")
        categorias:           dict[str, list[dict]]   nome canônico -> lançamentos
        totais_por_categoria: dict[str, float]
        total_geral:          float
        itens_fora_rateio:    list[dict]
    """
    condominio = ''
    periodo = ''
    lancamentos_raw: list[dict] = []

    with pdfplumber.open(caminho) as pdf:
        texto_completo = ''
        for pagina in pdf.pages:
            t = pagina.extract_text() or ''
            texto_completo += t + '\n'
            for linha in t.split('\n'):
                ls = linha.strip()
                if not ls:
                    continue
                if not condominio:
                    m = re.match(
                        r'condom[ií]nio[:\s]+(.+)$',
                        ls, re.IGNORECASE,
                    )
                    if m:
                        condominio = m.group(1).strip()
                if not periodo:
                    m = _RE_PERIODO.search(ls)
                    if m:
                        periodo = f'{m.group(1)} a {m.group(2)}'
                # Heurística: linha de lançamento tem data + descrição + valor BRL no fim
                m_data = _RE_DATA_BR.search(ls)
                m_valor = _RE_VALOR_BRL.search(ls)
                if m_data and m_valor and m_data.start() < m_valor.start():
                    descricao = ls[m_data.end():m_valor.start()].strip()
                    # Tira o "R$" residual e outros lixos do fim da descrição.
                    descricao = re.sub(r'\s*(?:R\$|\$)\s*$', '', descricao).strip()
                    if descricao:
                        valor = _parse_valor(m_valor.group(1))
                        if valor is not None:
                            lancamentos_raw.append({
                                'data': m_data.group(1),
                                'descricao': descricao,
                                'valor': valor,
                            })

    # Agrupamento por categoria canônica, separando itens fora do rateio.
    categorias: dict[str, list[dict]] = defaultdict(list)
    fora_rateio: list[dict] = []
    for l in lancamentos_raw:
        if _eh_fora_rateio(l['descricao']):
            fora_rateio.append(l)
            continue
        cat = _classificar_categoria(l['descricao'])
        categorias[cat].append(l)

    totais_por_categoria = {
        cat: round(sum(l['valor'] for l in lst), 2)
        for cat, lst in categorias.items()
    }
    total_geral = round(
        sum(v for v in totais_por_categoria.values())
        + sum(l['valor'] for l in fora_rateio),
        2,
    )

    # Aviso de cobertura: itens em 'Outros' precisam ser mapeados.
    if categorias.get('Outros'):
        descs = ', '.join(
            sorted({l['descricao'] for l in categorias['Outros']})[:5]
        )
        print(
            f'[parser] {len(categorias["Outros"])} lançamentos não mapeados '
            f'(amostra: {descs}). Atualize CATEGORIA_KEYWORDS.',
            file=sys.stderr,
        )

    return {
        'condominio': condominio or '[a confirmar]',
        'periodo': periodo or '[a confirmar]',
        'categorias': dict(categorias),
        'totais_por_categoria': totais_por_categoria,
        'total_geral': total_geral,
        'itens_fora_rateio': fora_rateio,
    }


# ─── W045A ───────────────────────────────────────────────────────────────

_RE_FRACAO = re.compile(r'(\d+[.,]\d{3,8})')


def parsear_w045a(caminho: str) -> list[dict[str, Any]]:
    """Extrai pares (unidade, fração) do W045A.

    Valida que a soma fecha em 1.0 com tolerância 0.001. Se falhar, loga
    warning no stderr mas devolve a lista mesmo assim para inspeção.
    """
    unidades: list[dict] = []
    with pdfplumber.open(caminho) as pdf:
        for pagina in pdf.pages:
            for linha in (pagina.extract_text() or '').split('\n'):
                ls = linha.strip()
                if not ls:
                    continue
                # Linha típica: "Apto 0101-1   0.004226"  ou  "Apto 0101-1   0,004226"
                m = _RE_FRACAO.search(ls)
                if not m:
                    continue
                fracao_str = m.group(1).replace(',', '.')
                try:
                    fracao = float(fracao_str)
                except ValueError:
                    continue
                # Filtro de sanidade: frações de unidade ficam entre 0.0001 e 0.5
                if not (0.0001 <= fracao <= 0.5):
                    continue
                unidade = ls[:m.start()].strip()
                if not unidade:
                    continue
                unidades.append({'unidade': unidade, 'fracao': fracao})

    soma = sum(u['fracao'] for u in unidades)
    if abs(soma - 1.0) > 0.001:
        print(
            f'[parser] soma das frações = {soma:.6f} (esperado 1.0 ± 0.001). '
            'Verifique o PDF ou ajuste o filtro em parsear_w045a.',
            file=sys.stderr,
        )
    return unidades


# ─── Geração da planilha ─────────────────────────────────────────────────

_FUNDO_RESERVA_PCT = 0.05  # 5%, padrão Naturale Residence
_FATOR_COBERTURA = 1.5

_FILL_HEADER = PatternFill('solid', fgColor='1F4E78')
_FILL_SUBTOTAL = PatternFill('solid', fgColor='D9E1F2')
_FILL_TOTAL = PatternFill('solid', fgColor='305496')
_FONT_HEADER = Font(bold=True, color='FFFFFF', size=12)
_FONT_SUBTOTAL = Font(bold=True, size=11)
_FONT_TOTAL = Font(bold=True, color='FFFFFF', size=12)
_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
_ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
_ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

_FMT_BRL = 'R$ #,##0.00'
_FMT_PCT = '0.0%'
_FMT_FRACAO = '0.000000'


def _set_header(ws, row: int, col: int, value: str) -> None:
    c = ws.cell(row, col, value)
    c.fill = _FILL_HEADER
    c.font = _FONT_HEADER
    c.alignment = _ALIGN_CENTER


def _aba_reajustes(wb: Workbook, dados: dict) -> None:
    ws = wb.create_sheet('Reajustes')
    ws.cell(2, 2, 'PAINEL DE REAJUSTE').font = Font(bold=True, size=14)
    ws.cell(3, 2, 'Edite os percentuais por categoria ou por item.')
    ws.cell(5, 2, "='Previsao Anual'!E122")
    ws.cell(5, 4, "='Resumo Assembleia'!D7")
    ws.cell(6, 2, 'Total Previsto')
    ws.cell(6, 4, 'Taxa Mensal (referência)')

    ws.cell(9, 2, 'REAJUSTE POR CATEGORIA').font = Font(bold=True, size=12)
    _set_header(ws, 10, 2, 'Categoria')
    _set_header(ws, 10, 3, 'Reajuste %')
    for i, cat in enumerate(CATEGORIAS_CANONICAS):
        ws.cell(11 + i, 2, cat)
        ws.cell(11 + i, 3, 0.0).number_format = _FMT_PCT

    ws.cell(20, 2, 'REAJUSTE POR ITEM').font = Font(bold=True, size=12)
    ws.cell(
        21, 2,
        'Preencha APENAS a coluna Reajuste % do item para sobrescrever a '
        'categoria.',
    )
    cabecalhos = [
        'Item', 'Categoria', 'Base Anual (R$)', 'Reajuste % do item',
        'Valor Anual 2026', 'Valor Mensal 2026', '% Aplicado',
        'Valor Final Anual 2026',
    ]
    for i, h in enumerate(cabecalhos):
        _set_header(ws, 22, 2 + i, h)

    linha = 23
    for cat in CATEGORIAS_CANONICAS:
        for lanc in dados['categorias'].get(cat, []):
            ws.cell(linha, 2, lanc['descricao'])
            ws.cell(linha, 3, cat)
            ws.cell(linha, 4, lanc['valor']).number_format = _FMT_BRL
            ws.cell(linha, 5, 0.0).number_format = _FMT_PCT
            ws.cell(linha, 6, lanc['valor']).number_format = _FMT_BRL
            ws.cell(linha, 7, lanc['valor'] / 12).number_format = _FMT_BRL
            linha += 1

    for col, width in zip(range(2, 10), (38, 28, 18, 16, 18, 18, 14, 20)):
        ws.column_dimensions[get_column_letter(col)].width = width


def _aba_previsao_anual(wb: Workbook, dados: dict) -> None:
    ws = wb.create_sheet('Previsao Anual')
    ws.cell(2, 2, f'PREVISÃO ANUAL 2026').font = Font(bold=True, size=14)
    ws.cell(
        3, 2,
        'Os percentuais vêm da aba Reajustes. Subtotais por categoria.',
    )
    cabecalhos = ['Item', 'Base Anual (R$)', 'Reajuste %',
                  'Previsão 2026 (R$)', 'Mensal (R$)']
    for i, h in enumerate(cabecalhos):
        _set_header(ws, 5, 2 + i, h)

    linha = 6
    for cat in CATEGORIAS_CANONICAS:
        # Cabeçalho da categoria
        c = ws.cell(linha, 2, cat)
        c.font = Font(bold=True)
        c.fill = _FILL_SUBTOTAL
        linha += 1
        for lanc in dados['categorias'].get(cat, []):
            ws.cell(linha, 2, lanc['descricao'])
            ws.cell(linha, 3, lanc['valor']).number_format = _FMT_BRL
            ws.cell(linha, 4, 0.0).number_format = _FMT_PCT
            ws.cell(linha, 5, lanc['valor']).number_format = _FMT_BRL
            ws.cell(linha, 6, lanc['valor'] / 12).number_format = _FMT_BRL
            linha += 1
        # Subtotal da categoria
        sub = dados['totais_por_categoria'].get(cat, 0.0)
        c = ws.cell(linha, 2, f'  Total {cat}')
        c.font = _FONT_SUBTOTAL
        c.fill = _FILL_SUBTOTAL
        ws.cell(linha, 5, sub).number_format = _FMT_BRL
        ws.cell(linha, 6, sub / 12).number_format = _FMT_BRL
        linha += 2

    total = sum(dados['totais_por_categoria'].values())
    c = ws.cell(linha, 2, 'TOTAL GERAL DE DESPESAS')
    c.font = _FONT_TOTAL
    c.fill = _FILL_TOTAL
    c = ws.cell(linha, 5, total)
    c.number_format = _FMT_BRL
    c.font = _FONT_TOTAL
    c.fill = _FILL_TOTAL
    c = ws.cell(linha, 6, total / 12)
    c.number_format = _FMT_BRL
    c.font = _FONT_TOTAL
    c.fill = _FILL_TOTAL

    for col, width in zip(range(2, 7), (40, 20, 14, 22, 18)):
        ws.column_dimensions[get_column_letter(col)].width = width


def _aba_previsao_mensal(wb: Workbook, dados: dict) -> None:
    ws = wb.create_sheet('Previsao Mensal')
    ws.cell(2, 2, 'PREVISÃO MENSAL 2026').font = Font(bold=True, size=14)
    ws.cell(3, 2, 'Valores mensais já reajustados.')
    meses = ['Jan/26', 'Fev/26', 'Mar/26', 'Abr/26', 'Mai/26', 'Jun/26',
             'Jul/26', 'Ago/26', 'Set/26', 'Out/26', 'Nov/26', 'Dez/26']
    _set_header(ws, 5, 2, 'Item')
    for i, m in enumerate(meses):
        _set_header(ws, 5, 3 + i, m)
    _set_header(ws, 5, 15, 'Total')

    linha = 6
    for cat in CATEGORIAS_CANONICAS:
        c = ws.cell(linha, 2, cat)
        c.font = Font(bold=True)
        c.fill = _FILL_SUBTOTAL
        linha += 1
        for lanc in dados['categorias'].get(cat, []):
            ws.cell(linha, 2, lanc['descricao'])
            mensal = lanc['valor'] / 12
            for i in range(12):
                ws.cell(linha, 3 + i, mensal).number_format = _FMT_BRL
            ws.cell(linha, 15, lanc['valor']).number_format = _FMT_BRL
            linha += 1
        # Subtotal mensal da categoria
        sub = dados['totais_por_categoria'].get(cat, 0.0)
        c = ws.cell(linha, 2, f'  Total {cat}')
        c.font = _FONT_SUBTOTAL
        c.fill = _FILL_SUBTOTAL
        for i in range(12):
            ws.cell(linha, 3 + i, sub / 12).number_format = _FMT_BRL
        ws.cell(linha, 15, sub).number_format = _FMT_BRL
        linha += 2

    total = sum(dados['totais_por_categoria'].values())
    c = ws.cell(linha, 2, 'TOTAL GERAL DE DESPESAS')
    c.font = _FONT_TOTAL
    c.fill = _FILL_TOTAL
    for i in range(12):
        c = ws.cell(linha, 3 + i, total / 12)
        c.number_format = _FMT_BRL
        c.font = _FONT_TOTAL
        c.fill = _FILL_TOTAL
    c = ws.cell(linha, 15, total)
    c.number_format = _FMT_BRL
    c.font = _FONT_TOTAL
    c.fill = _FILL_TOTAL

    ws.column_dimensions['B'].width = 40
    for i in range(12):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13
    ws.column_dimensions['O'].width = 16


def _aba_resumo_assembleia(wb: Workbook, dados_w011: dict,
                           dados_w045: list[dict]) -> None:
    ws = wb.create_sheet('Resumo Assembleia')
    ws.cell(2, 2, 'PREVISÃO ORÇAMENTÁRIA 2026').font = Font(bold=True, size=14)
    ws.cell(3, 2, f"Condomínio {dados_w011.get('condominio','[a confirmar]')}")

    ws.cell(5, 2, 'RESUMO FINANCEIRO 2026').font = Font(bold=True, size=12)

    total_desp = sum(dados_w011['totais_por_categoria'].values())
    fundo = round(total_desp * _FUNDO_RESERVA_PCT, 2)
    total_ratear = round(total_desp + fundo, 2)

    ws.cell(7, 2, 'Despesa Operacional Anual')
    ws.cell(7, 4, total_desp).number_format = _FMT_BRL
    ws.cell(10, 2, f'Fundo de Reserva ({int(_FUNDO_RESERVA_PCT * 100)}%)')
    ws.cell(10, 4, fundo).number_format = _FMT_BRL
    ws.cell(13, 2, 'Total a Ratear no Ano')
    c = ws.cell(13, 4, total_ratear)
    c.number_format = _FMT_BRL
    c.font = Font(bold=True)

    ws.cell(17, 2, 'MODO DE RATEIO').font = Font(bold=True, size=12)
    ws.cell(19, 2, 'Modo de cálculo:')
    ws.cell(19, 4, 'Fração Ideal' if dados_w045 else 'Igualdade')

    apartamentos = len(dados_w045)
    ws.cell(21, 2, 'Apartamentos:')
    ws.cell(21, 4, apartamentos)
    ws.cell(22, 2, 'Coberturas:')
    ws.cell(22, 4, 0)
    ws.cell(23, 2, 'Fator da Cobertura:')
    ws.cell(23, 4, _FATOR_COBERTURA)
    ws.cell(24, 2, 'Total de Unidades (igualdade):')
    ws.cell(24, 4, apartamentos)
    soma_fracao = round(sum(u['fracao'] for u in dados_w045), 6) if dados_w045 else 0.0
    ws.cell(25, 2, 'Soma das Frações (deve = 1.0):')
    ws.cell(25, 4, soma_fracao).number_format = _FMT_FRACAO

    ws.cell(29, 2, 'Despesa Mensal a Ratear:')
    ws.cell(29, 4, total_ratear / 12).number_format = _FMT_BRL

    ws.cell(31, 2, 'TAXA MENSAL RESULTANTE').font = Font(bold=True, size=12)
    taxa_apto_media = (
        total_ratear / 12 / apartamentos if apartamentos else 0.0
    )
    ws.cell(33, 2, 'Taxa Mensal Média por Apartamento:')
    ws.cell(33, 4, taxa_apto_media).number_format = _FMT_BRL
    ws.cell(34, 2, 'Taxa Mensal por Cobertura (fator 1,5):')
    ws.cell(34, 4, taxa_apto_media * _FATOR_COBERTURA).number_format = _FMT_BRL
    ws.cell(
        36, 2,
        'No modo Fração Ideal, cada unidade paga sua fração x Total a Ratear.',
    )

    ws.cell(40, 2, 'CUSTO MENSAL POR CATEGORIA').font = Font(bold=True, size=12)
    _set_header(ws, 41, 2, 'Categoria')
    _set_header(ws, 41, 3, 'Anual (R$)')
    _set_header(ws, 41, 4, 'Mensal (R$)')
    for i, cat in enumerate(CATEGORIAS_CANONICAS):
        anual = dados_w011['totais_por_categoria'].get(cat, 0.0)
        ws.cell(42 + i, 2, cat)
        ws.cell(42 + i, 3, anual).number_format = _FMT_BRL
        ws.cell(42 + i, 4, anual / 12).number_format = _FMT_BRL
    c = ws.cell(51, 2, 'TOTAL GERAL')
    c.font = _FONT_TOTAL
    c.fill = _FILL_TOTAL
    c = ws.cell(51, 3, total_desp)
    c.number_format = _FMT_BRL
    c.font = _FONT_TOTAL
    c.fill = _FILL_TOTAL
    c = ws.cell(51, 4, total_desp / 12)
    c.number_format = _FMT_BRL
    c.font = _FONT_TOTAL
    c.fill = _FILL_TOTAL

    for col, width in zip(range(2, 7), (34, 18, 16, 18, 16)):
        ws.column_dimensions[get_column_letter(col)].width = width


def _aba_fracoes(wb: Workbook, dados_w045: list[dict],
                 taxa_apto_media: float) -> None:
    ws = wb.create_sheet('Frações')
    _set_header(ws, 1, 2, 'Unidade')
    _set_header(ws, 1, 3, 'Fração')
    _set_header(ws, 1, 4, 'Taxa Mensal (R$)')
    for i, u in enumerate(dados_w045):
        ws.cell(2 + i, 2, u['unidade'])
        ws.cell(2 + i, 3, u['fracao']).number_format = _FMT_FRACAO
        # Taxa por fração ideal: fracao * total_ratear / 12. Em vez de fórmula,
        # gravamos o valor calculado para o XLSX abrir sem dependência cruzada.
        ws.cell(2 + i, 4, u['fracao'] * taxa_apto_media * len(dados_w045)).number_format = _FMT_BRL
    for col, width in zip(range(2, 5), (22, 14, 18)):
        ws.column_dimensions[get_column_letter(col)].width = width


def _aba_notas(wb: Workbook, dados_w011: dict) -> None:
    ws = wb.create_sheet('Notas (Por Fora)')
    ws.cell(2, 2, 'ITENS POR FORA DO RATEIO').font = Font(bold=True, size=14)
    ws.cell(
        3, 2,
        'Itens que NÃO entram na taxa condominial padrão. Rateados por uso '
        'real ou tratados como dívida específica.',
    )
    cabecalhos = ['Item', 'Categoria de Origem', 'Total Anual (R$)', 'Motivo']
    for i, h in enumerate(cabecalhos):
        _set_header(ws, 5, 2 + i, h)
    linha = 6
    for lanc in dados_w011.get('itens_fora_rateio', []):
        ws.cell(linha, 2, lanc['descricao'])
        ws.cell(linha, 3, _classificar_categoria(lanc['descricao']))
        ws.cell(linha, 4, lanc['valor']).number_format = _FMT_BRL
        ws.cell(linha, 5, 'Rateado separadamente / dívida específica')
        linha += 1
    total_fora = round(
        sum(l['valor'] for l in dados_w011.get('itens_fora_rateio', [])),
        2,
    )
    c = ws.cell(linha + 1, 2, 'TOTAL POR FORA')
    c.font = _FONT_TOTAL
    c.fill = _FILL_TOTAL
    c = ws.cell(linha + 1, 4, total_fora)
    c.number_format = _FMT_BRL
    c.font = _FONT_TOTAL
    c.fill = _FILL_TOTAL
    for col, width in zip(range(2, 6), (36, 26, 18, 38)):
        ws.column_dimensions[get_column_letter(col)].width = width


def gerar_planilha(dados_w011a: dict, dados_w045a: list[dict],
                   caminho_saida: str) -> None:
    """Monta o XLSX com as 6 abas no formato da referência Naturale_2026."""
    wb = Workbook()
    # Remove a aba default vazia
    wb.remove(wb.active)
    _aba_reajustes(wb, dados_w011a)
    _aba_previsao_anual(wb, dados_w011a)
    _aba_previsao_mensal(wb, dados_w011a)
    _aba_resumo_assembleia(wb, dados_w011a, dados_w045a)
    total_desp = sum(dados_w011a['totais_por_categoria'].values())
    fundo = total_desp * _FUNDO_RESERVA_PCT
    total_ratear = total_desp + fundo
    taxa_media = (
        total_ratear / 12 / len(dados_w045a) if dados_w045a else 0.0
    )
    _aba_fracoes(wb, dados_w045a, taxa_media)
    _aba_notas(wb, dados_w011a)
    wb.save(caminho_saida)


# ─── CLI ─────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--w011a', help='Caminho do PDF W011A (despesas).')
    parser.add_argument('--w045a', help='Caminho do PDF W045A (frações).')
    parser.add_argument(
        '--saida', default='previsao.xlsx',
        help='Caminho do XLSX de saída. Default: previsao.xlsx',
    )
    parser.add_argument(
        '--identificar', metavar='PDF',
        help='Apenas identifica o tipo do PDF (W011A/W045A/DESCONHECIDO).',
    )
    args = parser.parse_args()

    if args.identificar:
        print(identificar_tipo_pdf(args.identificar))
        return 0

    if not args.w011a:
        parser.error('--w011a é obrigatório (ou use --identificar).')
        return 2

    tipo = identificar_tipo_pdf(args.w011a)
    if tipo != 'W011A':
        print(
            f'[parser] aviso: {args.w011a} foi identificado como {tipo}. '
            'Prosseguindo mesmo assim.',
            file=sys.stderr,
        )
    dados_w011 = parsear_w011a(args.w011a)

    dados_w045: list[dict] = []
    if args.w045a:
        tipo45 = identificar_tipo_pdf(args.w045a)
        if tipo45 != 'W045A':
            print(
                f'[parser] aviso: {args.w045a} foi identificado como {tipo45}.',
                file=sys.stderr,
            )
        dados_w045 = parsear_w045a(args.w045a)

    gerar_planilha(dados_w011, dados_w045, args.saida)
    print(f'OK: {args.saida}')
    print(f'  condomínio: {dados_w011["condominio"]}')
    print(f'  período:    {dados_w011["periodo"]}')
    print(f'  total:      R$ {dados_w011["total_geral"]:,.2f}')
    print(f'  categorias com dado: {len(dados_w011["categorias"])}')
    print(f'  unidades W045A:      {len(dados_w045)}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
